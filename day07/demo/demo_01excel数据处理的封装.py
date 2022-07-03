"""
# @Time : 2022/5/24 0024   16:49
# @File : demo_01excel数据处理的封装
# @Project : python01
# @Content :
"""

"""
封装的需求：
    1、封装一个可以读取任意excel文件的方法，可以指定表单
    2、数据写入:
        文件名：
        表单：
        行：
        列：
        写入的值：
        
"""
import openpyxl


class HandleExcel:

    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    # def read_data(self, filename, sheetname):
    #     """读取excel方法"""
    #     workbook = openpyxl.load_workbook(filename)
    #     sh = workbook[sheetname]
    #     res = list(sh.rows)
    #     # 获取第一行的表头
    #     title = [i.value for i in res[0]]
    #
    #     cases = []
    #     # 遍历第一行以外的其他行
    #     for item in res[1:]:
    #         data = [i.value for i in item]
    #         dic = dict(zip(title, data))
    #         cases.append(dic)
    #     return cases

    def read_data(self):
        """读取excel方法"""
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        res = list(sh.rows)
        # 获取第一行的表头
        title = [i.value for i in res[0]]
        cases = []
        # 遍历第一行以外的其他行
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        return cases

    def write_data(self, row, column, value):
        """数据写入的方法"""
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        sh.cell(row=row, column=column, value=value)
        workbook.save(self.filename)


if __name__ == '__main__':
    # excel = HandleExcel()
    # res = excel.read_data(r'D:\pythonworkspace\python01\day07\cases.xlsx', 'yanzu')
    # print(res)
    excel = HandleExcel(r'D:\pythonworkspace\python01\day07\cases.xlsx', 'yanzu')
    res = excel.read_data()
    print(res)
