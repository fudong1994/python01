"""
# @Time : 2022/5/24 0024   17:27
# @File : handle_excel
# @Project : python01
# @Content :
"""

import openpyxl


class HandleExcel:

    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        """读取excel方法"""
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        res = list(sh.rows)
        # 获取第一行的表头
        title = [i.value for i in res[0]]
        cases = []
        # 遍历第一行以外的其他行
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        return cases

    def write_data(self, row, column, value):
        """数据写入的方法"""
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        sh.cell(row=row, column=column, value=value)
        workbook.save(self.filename)
