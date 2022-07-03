"""
# @Time : 2022/5/24 0024   14:41
# @File : __init__.py
# @Project : python01
# @Content :
"""

import openpyxl

# ----------------基础用法----------------
# 加载excel文件为工作蒲对象
# workbook = openpyxl.load_workbook("cases.xlsx")
# 获取所有的表单名
# print(workbook.sheetnames)

# 选中表单
# sh = workbook["Sheet1"]
# print(sh)

# 读取数据
# c = sh.cell(row=1, column=1)
# print(c.value)
# print(sh.cell(row=3, column=3).value)

# ----------------读取表单所有的数据----------------
workbook = openpyxl.load_workbook("case.xlsx")
sh = workbook["Sheet1"]

# rows:按行获取表单中所有的格子，每一行放到一个元组中
res = list(sh.rows)
print(res)

# rows:按列获取表单中所有的格子，每一列放到一个元组中
# res1 = list(sh.columns)
# print(res)
